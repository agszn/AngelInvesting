from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST

from django.views.decorators.csrf import csrf_exempt

from django.utils.timezone import localtime
from django.utils import timezone

from decimal import Decimal
from django.db.models import Q

from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse

import json
from collections import defaultdict

from .forms import *
from .models import *
from .utils import *
from .views import *

from unlisted_stock_marketplace.models import *

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import UserPortfolioSummary, BuyTransaction
from decimal import Decimal


from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render
from .models import UserPortfolioSummary, BuyTransactionOtherAdvisor

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Sum
from decimal import Decimal
from .models import UserPortfolioSummary, BuyTransactionOtherAdvisor

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from user_portfolio.models import UserStockInvestmentSummary
from user_portfolio.utils import get_user_stock_context


def _parse_sort(request, default='-timestamp'):
    sort = request.GET.get('sort')
    if sort == 'date_asc':
        return 'timestamp'
    elif sort == 'date_desc':
        return '-timestamp'
    return default

def _market_closed_note():
    now = timezone.localtime()
    if now.hour >= 18:
        return "The market is closed. Your order will be processed tomorrow at 11:00 AM."
    return None
@login_required
def profile_overview(request):
    user = request.user
    summaries = UserStockInvestmentSummary.objects.filter(user=user)

    def aggregate_by_stock_type(stock_type):
        filtered = summaries.filter(stock__stock_type=stock_type)
        return {
            "total_invested": sum((s.investment_amount for s in filtered), Decimal("0.00")),
            "market_value": sum((s.market_value for s in filtered), Decimal("0.00")),
            "overall_gain_loss": sum((s.overall_gain_loss for s in filtered), Decimal("0.00")),
            "todays_gain_loss": sum((s.day_gain_loss for s in filtered), Decimal("0.00")),
        }

    # Aggregated summaries
    unlisted_summary = aggregate_by_stock_type("Unlisted")
    angel_summary = aggregate_by_stock_type("Angel")

    # Combined totals
    total_invested = unlisted_summary["total_invested"] + angel_summary["total_invested"]
    total_market_value = unlisted_summary["market_value"] + angel_summary["market_value"]
    total_gain_loss = unlisted_summary["overall_gain_loss"] + angel_summary["overall_gain_loss"]
    todays_gain = unlisted_summary["todays_gain_loss"] + angel_summary["todays_gain_loss"]

    def percent(numerator, denominator):
        return round((numerator / denominator * 100), 2) if denominator else 0

    context = {
        # Totals for top summary
        "total_invested": total_invested,
        "total_market_value": total_market_value,
        "total_gain_loss": total_gain_loss,
        "todays_gain": todays_gain,
        "overall_gain_loss_percent": percent(total_gain_loss, total_invested),
        "todays_gain_percent": percent(todays_gain, total_invested),

        # Individual summaries for breakdown rows
        "unlisted_summary": unlisted_summary,
        "unlisted_overall_percent": percent(unlisted_summary["overall_gain_loss"], unlisted_summary["total_invested"]),
        "unlisted_today_percent": percent(unlisted_summary["todays_gain_loss"], unlisted_summary["total_invested"]),

        "angel_summary": angel_summary,
        "angel_overall_percent": percent(angel_summary["overall_gain_loss"], angel_summary["total_invested"]),
        "angel_today_percent": percent(angel_summary["todays_gain_loss"], angel_summary["total_invested"]),

        **get_user_stock_context(user, request),
    }

    return render(request, "portfolio/overview.html", context)

from django.contrib import messages

from django.db.models import Sum, F
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserPortfolioSummary, UserStockInvestmentSummary
from .utils import get_user_stock_context
from decimal import Decimal

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

from user_portfolio.models import UserPortfolioSummary, BuyTransactionOtherAdvisor
from user_portfolio.utils import get_user_stock_context

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

from user_portfolio.models import UserPortfolioSummary, BuyTransactionOtherAdvisor
from user_portfolio.utils import get_user_stock_context

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from user_portfolio.models import UserStockInvestmentSummary
from user_portfolio.utils import get_user_stock_context

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from user_portfolio.models import UserStockInvestmentSummary

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from user_portfolio.models import UserStockInvestmentSummary

@login_required
def unlisted_view(request):
    user = request.user
    unlisted_all = UserStockInvestmentSummary.objects.filter(user=user, stock__stock_type="Unlisted")

    # Split based on is_other_advisor flag
    thangiv_qs = unlisted_all.filter(is_other_advisor=False)
    other_qs = unlisted_all.filter(is_other_advisor=True)

    def get_summary(qs):
        return {
            "total_invested": sum((s.investment_amount for s in qs), Decimal("0.00")),
            "total_market_value": sum((s.market_value for s in qs), Decimal("0.00")),
            "overall_gain_loss": sum((s.overall_gain_loss for s in qs), Decimal("0.00")),
            "todays_gain_loss": sum((s.day_gain_loss for s in qs), Decimal("0.00")),
        }

    thangiv_summary = get_summary(thangiv_qs)
    other_summary = get_summary(other_qs)

    # Combine for top box
    total_invested = thangiv_summary["total_invested"] + other_summary["total_invested"]
    total_market_value = thangiv_summary["total_market_value"] + other_summary["total_market_value"]
    total_gain_loss = thangiv_summary["overall_gain_loss"] + other_summary["overall_gain_loss"]
    todays_gain_loss = thangiv_summary["todays_gain_loss"] + other_summary["todays_gain_loss"]

    def percent(numerator, denominator):
        return round((numerator / denominator * 100), 2) if denominator else 0

    context = {
        # Top summary box
        "total_invested": total_invested,
        "total_market_value": total_market_value,
        "total_gain_loss": total_gain_loss,
        "todays_gain": todays_gain_loss,
        "overall_gain_loss_percent": percent(total_gain_loss, total_invested),
        "todays_gain_percent": percent(todays_gain_loss, total_invested),

        # Thangiv (non-other)
        "portfolio_summary": thangiv_summary,
        "unlisted_overall_percent": percent(thangiv_summary["overall_gain_loss"], thangiv_summary["total_invested"]),
        "unlisted_today_percent": percent(thangiv_summary["todays_gain_loss"], thangiv_summary["total_invested"]),

        # Other
        "advisor_summary": other_summary,
        "advisor_overall_percent": percent(other_summary["overall_gain_loss"], other_summary["total_invested"]),
        "advisor_today_percent": percent(other_summary["todays_gain_loss"], other_summary["total_invested"]),

        **get_user_stock_context(user, request),
    }

    return render(request, 'portfolio/unlistedOverview.html', context)

@login_required
def angel_invest(request):
    user = request.user
    stock_context = get_user_stock_context(user, request)
    
    return render(request, 'portfolio/angelOverview.html',{ **stock_context})

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import *
from .utils import update_user_holdings

from django.core.paginator import Paginator
from django.db.models import Q
from site_Manager.models import *
from django.utils.timezone import make_aware
from datetime import datetime, timezone
from user_portfolio.models import UserStockInvestmentSummary


# this is portfolio with other advisor in different table
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from .models import UserStockInvestmentSummary, Advisor, Broker
from .utils import update_user_holdings  # make sure this is in utils.py
from io import BytesIO
from decimal import Decimal
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers

from site_Manager.models import Advisor, Broker
from .models import UserStockInvestmentSummary
from .utils import update_user_holdings
# If you have this helper in your codebase:
# from .helpers import get_user_stock_context

# views.py

from io import BytesIO
from decimal import Decimal
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers

from site_Manager.models import Advisor, Broker
from .models import UserStockInvestmentSummary
from .utils import update_user_holdings
# If you have this helper in your codebase:
# from .helpers import get_user_stock_context

# views.py
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

from .models import UserStockInvestmentSummary, Advisor, Broker  # adjust paths if needed
from .utils import update_user_holdings  # adjust import if defined elsewhere
# from .somewhere import get_user_stock_context  # only if you have it; handled defensively below
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import render
from django.http import HttpResponse
from django.core.paginator import Paginator
from decimal import Decimal
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from .models import UserStockInvestmentSummary, Advisor, Broker
from .utils import update_user_holdings, get_user_stock_context

@login_required
def portfolio_view(request):
    user = request.user
    update_user_holdings(user)

    qs = (
        UserStockInvestmentSummary.objects
        .filter(user=user)
        .select_related('stock', 'advisor', 'broker')
    )

    # ---- Filters & sorting from querystring ----
    advisor_param = request.GET.get('advisor')  # None | 'all' | id | 'THANGIV'/'OTHER'/...
    broker_param  = request.GET.get('broker')   # None | 'all' | id
    search        = (request.GET.get('search') or '').strip()
    sort          = (request.GET.get('sort') or '').lower()  # 'inv' | 'mkt' | ''
    dir_          = (request.GET.get('dir') or 'desc').lower()
    if dir_ not in ('asc', 'desc'):
        dir_ = 'desc'

    # Detect if 'advisor' key was present at all (important for defaulting logic)
    advisor_key_present = 'advisor' in request.GET

    selected_advisor = None
    selected_broker  = None
    advisor_effective = None  # 'default' or 'forced' when we set it implicitly

    # ----- Advisor selection / defaulting -----
    if advisor_param and str(advisor_param).lower() != 'all':
        # Explicit advisor filter in query
        if str(advisor_param).isdigit():
            qs = qs.filter(advisor__id=int(advisor_param))
            selected_advisor = Advisor.objects.filter(id=advisor_param).first()
        else:
            qs = qs.filter(advisor__advisor_type__iexact=advisor_param)
            selected_advisor = Advisor.objects.filter(advisor_type__iexact=advisor_param).first()
    else:
        # No explicit advisor (either not present OR 'all')
        if not advisor_key_present and not broker_param and not search:
            # Clean landing → default to Thangiv
            default_adv = Advisor.objects.filter(advisor_type__iexact='Thangiv').first()
            if default_adv:
                qs = qs.filter(advisor=default_adv)
                selected_advisor = default_adv
                advisor_effective = 'default'
        elif (  # Broker-only filter → force Thangiv so label matches data
            (advisor_param is None or str(advisor_param).lower() == 'all')
            and broker_param and str(broker_param).lower() != 'all'
            and not search
        ):
            forced_adv = Advisor.objects.filter(advisor_type__iexact='Thangiv').first()
            if forced_adv:
                qs = qs.filter(advisor=forced_adv)
                selected_advisor = forced_adv
                advisor_effective = 'forced'

    # ----- Broker filter (narrows qs). Merge ignores broker but filter must respect it. -----
    if broker_param and str(broker_param).lower() != 'all':
        if str(broker_param).isdigit():
            qs = qs.filter(broker__id=int(broker_param))
            selected_broker = Broker.objects.filter(id=broker_param).first()
        # (extend for name-based filtering if needed)

    # ----- Search -----
    if search:
        qs = qs.filter(
            Q(stock__company_name__icontains=search) |
            Q(stock__scrip_name__icontains=search)
        )

    # Hide zero-quantity rows
    qs = qs.exclude(quantity_held=0)

    # ---- Merge rows by (stock, advisor), preserve broker from latest row ----
    buckets = {}
    for row in qs:
        key = (row.stock_id, row.advisor_id)  # broker ignored in merge key
        agg = buckets.get(key)
        if not agg:
            agg = {
                'stock': row.stock,
                'advisor': row.advisor,
                'broker': row.broker,  # Preserve broker (latest seen below)
                'quantity_held': 0,
                'investment_amount': Decimal('0'),
                'market_value': Decimal('0'),
                'overall_gain_loss': Decimal('0'),
                'day_gain_loss': Decimal('0'),
                'buy_order_count': 0,
                'buy_order_total': Decimal('0'),
                'sell_order_count': 0,
                'sell_order_total': Decimal('0'),
                'share_price': row.share_price,
                'previous_day_price': row.previous_day_price,
                'last_updated': row.last_updated,
            }
            buckets[key] = agg

        agg['quantity_held']     += int(row.quantity_held or 0)
        agg['investment_amount'] += row.investment_amount or Decimal('0')
        agg['market_value']      += row.market_value or Decimal('0')
        agg['overall_gain_loss'] += row.overall_gain_loss or Decimal('0')
        agg['day_gain_loss']     += row.day_gain_loss or Decimal('0')
        agg['buy_order_count']   += int(row.buy_order_count or 0)
        agg['buy_order_total']   += row.buy_order_total or Decimal('0')
        agg['sell_order_count']  += int(row.sell_order_count or 0)
        agg['sell_order_total']  += row.sell_order_total or Decimal('0')

        # Update broker and prices based on latest row
        if row.last_updated and (not agg['last_updated'] or row.last_updated > agg['last_updated']):
            agg['last_updated'] = row.last_updated
            agg['broker'] = row.broker  # Update to latest broker
            if row.share_price is not None:
                agg['share_price'] = row.share_price
            if row.previous_day_price is not None:
                agg['previous_day_price'] = row.previous_day_price

    class Row: ...
    merged = []
    for a in buckets.values():
        qty = a['quantity_held']
        inv = a['investment_amount'] or Decimal('0')
        avg_price = (inv / qty) if qty else Decimal('0')
        overall_gain_percent = (a['overall_gain_loss'] / inv * 100) if inv > 0 else Decimal('0')
        day_gain_percent     = (a['day_gain_loss'] / inv * 100) if inv > 0 else Decimal('0')

        r = Row()
        r.stock                = a['stock']
        r.advisor              = a['advisor']
        r.broker               = a['broker']
        r.quantity_held        = qty
        r.avg_price            = avg_price
        r.share_price          = a['share_price']
        r.previous_day_price   = a['previous_day_price']
        r.market_value         = a['market_value']
        r.investment_amount    = inv
        r.overall_gain_loss    = a['overall_gain_loss']
        r.overall_gain_percent = overall_gain_percent
        r.day_gain_loss        = a['day_gain_loss']
        r.day_gain_percent     = day_gain_percent
        r.last_updated         = a['last_updated']
        merged.append(r)

    # Apply sorting: user-specified sort takes precedence over default
    if sort in ('inv', 'mkt'):
        if sort == 'inv':
            merged.sort(key=lambda x: x.investment_amount or Decimal('0'), reverse=(dir_ == 'desc'))
        else:
            merged.sort(key=lambda x: x.market_value or Decimal('0'), reverse=(dir_ == 'desc'))
    else:
        merged.sort(
            key=lambda x: (x.last_updated or datetime.min, x.stock.id if x.stock else 0),
            reverse=True
        )

    # ---------- XLSX ----------
    if (request.GET.get('download') or '').lower() == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio"
        headers = [
            "Stock Name", "Advisor", "Broker", "CMP", "Avg. Price", "Quantity",
            "Investment Amount", "Market Value", "Overall P/L", "Overall P/L %",
            "Day P/L", "Day P/L %", "Last Updated"
        ]
        ws.append(headers)
        for r in merged:
            overall_pct_fraction = float((r.overall_gain_percent or 0) / 100)
            day_pct_fraction     = float((r.day_gain_percent or 0) / 100)
            ws.append([
                (r.stock.company_name if r.stock else "-"),
                (r.advisor.advisor_type if r.advisor else "-"),
                (r.broker.name if r.broker else "-"),
                float(r.share_price or 0),
                float(r.avg_price or 0),
                int(r.quantity_held or 0),
                float(r.investment_amount or 0),
                float(r.market_value or 0),
                float(r.overall_gain_loss or 0),
                overall_pct_fraction,
                float(r.day_gain_loss or 0),
                day_pct_fraction,
                r.last_updated.isoformat(sep=" ", timespec="seconds") if r.last_updated else ""
            ])

        # Bold header, freeze, filter
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Number formats (now that Broker column exists, numeric block starts at col 4)
        for row_cells in ws.iter_rows(min_row=2, min_col=4, max_col=12):
            row_cells[0].number_format = numbers.FORMAT_NUMBER_00   # CMP
            row_cells[1].number_format = numbers.FORMAT_NUMBER_00   # Avg. Price
            row_cells[2].number_format = numbers.FORMAT_NUMBER      # Quantity
            row_cells[3].number_format = numbers.FORMAT_NUMBER_00   # Investment Amount
            row_cells[4].number_format = numbers.FORMAT_NUMBER_00   # Market Value
            row_cells[5].number_format = numbers.FORMAT_NUMBER_00   # Overall P/L
            row_cells[6].number_format = "0.00%"                    # Overall P/L %
            row_cells[7].number_format = numbers.FORMAT_NUMBER_00   # Day P/L
            row_cells[8].number_format = "0.00%"                    # Day P/L %

        # Column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 42)

        bio = BytesIO()
        wb.save(bio); bio.seek(0)
        resp = HttpResponse(
            bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp['Content-Disposition'] = 'attachment; filename="portfolio.xlsx"'
        return resp
    # ---------- /XLSX ----------

    paginator = Paginator(merged, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    stock_context = {}
    if 'get_user_stock_context' in globals():
        stock_context = get_user_stock_context(user, request)

    # ---- Labels & stable query values for template ----
    if advisor_effective in ('default', 'forced') and selected_advisor:
        advisor_label = (selected_advisor.advisor_type or "Advisor").upper()
        advisor_query_value = str(selected_advisor.id)  # reflect the effective advisor we applied
    else:
        if advisor_param is None:
            advisor_label = "ALL ADVISORS"
            advisor_query_value = "all"
        elif str(advisor_param).lower() == 'all':
            advisor_label = "ALL ADVISORS"
            advisor_query_value = "all"
        elif selected_advisor:
            advisor_label = (selected_advisor.advisor_type or "Advisor").upper()
            advisor_query_value = str(selected_advisor.id) if str(advisor_param).isdigit() else (selected_advisor.advisor_type or "")
        else:
            advisor_label = "ADVISOR"
            advisor_query_value = "all"

    broker_label = selected_broker.name if selected_broker else "All Brokers"
    broker_query_value = str(selected_broker.id) if selected_broker else "all"

    context = {
        'holdings': page_obj,
        'advisors': Advisor.objects.all(),
        'brokers': Broker.objects.all(),
        'selected_advisor': selected_advisor,
        'selected_broker': selected_broker,
        'advisor_param': advisor_param,
        'advisor_label': advisor_label,
        'advisor_query_value': advisor_query_value,
        'broker_query_value': broker_query_value,
        'search_query': search,
        'page_obj': page_obj,
        'sort': sort,
        'dir': dir_,
        **stock_context,
    }

    tpl = (
        'portfolio/includes/portfolio_table.html'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        else 'portfolio/PortfolioList.html'
    )
    return render(request, tpl, context)


from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render
from user_portfolio.models import BuyTransaction, BuyTransactionOtherAdvisor
from user_portfolio.utils import get_user_stock_context  # Assuming this exists

from itertools import chain
from operator import attrgetter
from django.utils.timezone import make_aware

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.core.paginator import Paginator
from itertools import chain
from operator import attrgetter
from .models import BuyTransaction, BuyTransactionOtherAdvisor, Advisor, Broker
from RM_User.models import *

# views.py
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden

from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden




from itertools import chain
from operator import attrgetter

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_POST

# import your helper if you have it
# from .utils import _parse_sort
# from .helpers import get_user_stock_context
# If those live elsewhere, keep as-is.
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.shortcuts import render
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from itertools import chain
from operator import attrgetter

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from django.utils import timezone
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
# imports you already have:
# from .utils import get_user_stock_context, _parse_sort


@login_required
def buy_orders(request):
    user       = request.user
    search     = (request.GET.get('search') or '').strip()
    advisor_id = request.GET.get('advisor')
    broker_id  = request.GET.get('broker')
    ordering   = _parse_sort(request, default='-timestamp')  # keep your helper

    qs1 = BuyTransaction.objects.filter(user=user).select_related('stock', 'advisor', 'broker')
    qs2 = BuyTransactionOtherAdvisor.objects.filter(user=user).select_related('stock', 'advisor', 'broker')

    # ---- Filters ----
    if advisor_id and advisor_id != 'all':
        qs1 = qs1.filter(advisor__id=advisor_id)
        qs2 = qs2.filter(advisor__id=advisor_id)

    if broker_id and broker_id != 'all':
        qs1 = qs1.filter(broker__id=broker_id)
        qs2 = qs2.filter(broker__id=broker_id)

    if search:
        for_q = Q(stock__company_name__icontains=search) | Q(stock__scrip_name__icontains=search)
        qs1 = qs1.filter(for_q)
        qs2 = qs2.filter(for_q)

    # ---- Merge + sort (Python-side because models differ) ----
    # ordering is '-timestamp' or 'timestamp' (your helper guarantees)
    reverse = ordering.startswith('-')
    orders = sorted(chain(qs1, qs2), key=attrgetter('timestamp'), reverse=reverse)

    # ---- Paginate ----
    paginator = Paginator(orders, 15)
    page_obj  = paginator.get_page(request.GET.get('page'))

    # ---- can_delete computation on visible page ----
    page_ids = [o.order_id for o in page_obj if getattr(o, 'order_id', None)]
    paid_ids = set(
        RMPaymentRecord.objects.filter(rm_user_view__order_id__in=page_ids)
        .values_list('rm_user_view__order_id', flat=True)
    )
    for o in page_obj:
        o.can_delete = (getattr(o, 'RM_status', None) == 'processing') and (o.order_id not in paid_ids)

    # ---- Additional context ----
    stock_context = get_user_stock_context(user, request)

    context = {
        'orders'          : page_obj,  # iterable page
        'page_obj'        : page_obj,
        'paginator'       : paginator,
        'advisors'        : Advisor.objects.all().order_by('advisor_type'),
        'brokers'         : Broker.objects.all().order_by('name'),
        'selected_advisor': Advisor.objects.filter(id=advisor_id).first() if advisor_id and advisor_id != 'all' else None,
        'selected_broker' : Broker.objects.filter(id=broker_id).first() if broker_id and broker_id != 'all' else None,
        'sort'            : ordering,
        'search_query'    : search,
        'now'             : timezone.localtime(),
        **stock_context,
    }

    tpl = (
        'portfolio/includes/buy_orders_table.html'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        else 'portfolio/BuyOrdersList.html'
    )
    return render(request, tpl, context)

@login_required
@require_POST
def delete_buy_order(request):
    """
    Deletes a buy order ONLY if:
      - it belongs to the requesting user
      - RM_status == 'processing'
      - there are NO RMPaymentRecord rows linked to this order_id
    Works for both BuyTransaction and BuyTransactionOtherAdvisor.
    """
    order_id = request.POST.get('order_id')
    if not order_id:
        return HttpResponseBadRequest('Missing order_id')

    obj = (
        BuyTransaction.objects.filter(order_id=order_id, user=request.user).first() or
        BuyTransactionOtherAdvisor.objects.filter(order_id=order_id, user=request.user).first()
    )
    if not obj:
        return HttpResponseForbidden('Order not found or not yours')

    if getattr(obj, 'RM_status', None) != 'processing':
        return JsonResponse({'ok': False, 'reason': 'locked_status'}, status=400)

    if RMPaymentRecord.objects.filter(rm_user_view__order_id=order_id).exists():
        return JsonResponse({'ok': False, 'reason': 'has_payment'}, status=400)

    obj.delete()
    return JsonResponse({'ok': True})



from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.core.paginator import Paginator
from .models import SellTransaction, Advisor, Broker
from .utils import get_user_stock_context  # assuming you have this
from itertools import chain
from operator import attrgetter
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import SellTransaction, SellTransactionOtherAdvisor, Advisor, Broker
from .utils import get_user_stock_context

from itertools import chain
from operator import attrgetter

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_POST

from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.db import transaction

@login_required
@require_POST
def delete_sell_order(request):
    """
    Deletes a sell order ONLY if:
      - it belongs to the requesting user
      - RM_status == 'processing'
      - there are NO RMPaymentRecord rows linked to this order_id
    Works for both SellTransaction and SellTransactionOtherAdvisor.
    """
    order_id = request.POST.get('order_id')
    if not order_id:
        return HttpResponseBadRequest('Missing order_id')

    obj = (
        SellTransaction.objects.filter(order_id=order_id, user=request.user).first()
        or SellTransactionOtherAdvisor.objects.filter(order_id=order_id, user=request.user).first()
    )
    if not obj:
        return HttpResponseForbidden('Order not found or not yours')

    # Normalize RM_status just in case
    rm_status = (getattr(obj, "RM_status", "") or "").lower()
    if rm_status != "processing":
        return JsonResponse({"ok": False, "reason": "locked_status"}, status=400)

    has_payment = RMPaymentRecord.objects.filter(
        rm_user_view__order_id=order_id
    ).exists()
    if has_payment:
        return JsonResponse({"ok": False, "reason": "has_payment"}, status=400)

    with transaction.atomic():
        obj.delete()

    return JsonResponse({"ok": True})

@login_required
def sell_orders(request):
    user = request.user
    search     = (request.GET.get('search') or '').strip()
    advisor_id = request.GET.get('advisor')
    broker_id  = request.GET.get('broker')
    ordering   = _parse_sort(request, default='-timestamp')
    status_filter = request.GET.get('status')

    qs1 = SellTransaction.objects.filter(user=user).select_related('stock', 'advisor', 'broker')
    qs2 = SellTransactionOtherAdvisor.objects.filter(user=user).select_related('stock', 'advisor', 'broker')

    if status_filter:
        qs1 = qs1.filter(status=status_filter)
        qs2 = qs2.filter(status=status_filter)

    if advisor_id and advisor_id != 'all':
        qs1 = qs1.filter(advisor__id=advisor_id)
        qs2 = qs2.filter(advisor__id=advisor_id)
    if broker_id and broker_id != 'all':
        qs1 = qs1.filter(broker__id=broker_id)
        qs2 = qs2.filter(broker__id=broker_id)
    if search:
        for_q = Q(stock__company_name__icontains=search) | Q(stock__scrip_name__icontains=search)
        qs1 = qs1.filter(for_q)
        qs2 = qs2.filter(for_q)

    orders = sorted(chain(qs1, qs2), key=attrgetter('timestamp'), reverse=ordering.startswith('-'))

    paginator = Paginator(orders, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Mark deletable for rows on this page:
    # can delete iff RM_status == 'processing' and NO RMPaymentRecord exists for this order_id
    page_ids = [o.order_id for o in page_obj if getattr(o, "order_id", None)]
    paid_ids = set(
        RMPaymentRecord.objects.filter(rm_user_view__order_id__in=page_ids)
        .values_list('rm_user_view__order_id', flat=True)
    )
    for o in page_obj:
        o.can_delete = (getattr(o, "RM_status", None) == "processing") and (o.order_id not in paid_ids)

    stock_context = get_user_stock_context(user, request)
    context = {
        'orders': page_obj,
        'advisors': Advisor.objects.all(),
        'brokers': Broker.objects.all(),
        'selected_advisor': Advisor.objects.filter(id=advisor_id).first() if advisor_id and advisor_id != 'all' else None,
        'selected_broker': Broker.objects.filter(id=broker_id).first() if broker_id and broker_id != 'all' else None,
        'sort': ordering,
        'search_query': search,
        'page_obj': page_obj,
        'now': timezone.localtime(),
        **stock_context,
    }
    tpl = 'portfolio/includes/sell_orders_table.html' if request.headers.get('X-Requested-With')=='XMLHttpRequest' else 'portfolio/SellOrdersList.html'
    return render(request, tpl, context)


# testing

# 
# 
# ---------------------------------------------------------------------
# ------------------ Buy stocks -------------
# ------------------------------------------------------------------
# 
# 
# user_portfolio/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import BuyTransaction, SellTransaction
from unlisted_stock_marketplace.models import *
from site_Manager.models import *
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from decimal import Decimal

from .models import BuyTransaction
from site_Manager.models import Broker, Advisor
from unlisted_stock_marketplace.models import StockData

from django.http import JsonResponse


def load_advisors_brokers(request):
    advisors = list(Advisor.objects.values('id', 'advisor_type'))
    brokers = list(Broker.objects.values('id', 'name'))
    return JsonResponse({
        'advisors': advisors,
        'brokers': brokers
    })


from .models import BuyTransaction, BuyTransactionOtherAdvisor
from user_auth.decorators import require_complete_profile

@require_POST
@login_required
def buy_stock(request, stock_id):
    stock = get_object_or_404(StockData, id=stock_id)

    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER', '/'))

    try:
        advisor = get_object_or_404(Advisor, id=request.POST.get('advisor'))
        broker = get_object_or_404(Broker, id=request.POST.get('broker'))
        quantity = int(request.POST.get('quantity'))
        order_type = request.POST.get('order_type', 'market')

        if order_type == 'market':
            price_per_share = stock.share_price or Decimal('0')
        else:
            price_per_share = Decimal(request.POST.get('price_per_share'))

        total_amount = price_per_share * quantity

        if advisor.advisor_type.lower() == 'other':
            BuyTransactionOtherAdvisor.objects.create(
                user=request.user, stock=stock, advisor=advisor, broker=broker,
                quantity=quantity, price_per_share=price_per_share,
                order_type=order_type, total_amount=total_amount, status='completed'
            )
        else:
            BuyTransaction.objects.create(
                user=request.user, stock=stock, advisor=advisor, broker=broker,
                quantity=quantity, price_per_share=price_per_share,
                order_type=order_type, total_amount=total_amount, status='processing'
            )

        note = _market_closed_note()
        if note:
            messages.info(request, note)
        messages.success(request, f"Buy order placed for {stock.company_name}.")

    except Exception as e:
        messages.error(request, f"Failed to place order: {e}")

    return redirect(request.META.get('HTTP_REFERER', '/'))


# 
# 
# ---------------------------------------------------------------------
# ------------------ Sell stocks -------------
# ------------------------------------------------------------------
# 
# 
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from .models import StockData, SellTransaction, Advisor, Broker
from decimal import Decimal, InvalidOperation

from decimal import Decimal, InvalidOperation
from .models import SellTransaction, SellTransactionOtherAdvisor, UserStockInvestmentSummary

from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from .models import StockData, SellTransaction, Advisor, Broker, UserStockInvestmentSummary
# at the imports near sell_stock
from django.db.models import Sum  # add this


@require_POST
@login_required
def sell_stock(request, stock_id):
    stock = get_object_or_404(StockData, id=stock_id)

    # (Redundant because of @require_POST, but keeping your original guard)
    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER', '/'))

    try:
        advisor = get_object_or_404(Advisor, id=request.POST.get('advisor'))
        broker = get_object_or_404(Broker, id=request.POST.get('broker'))

        # Parse quantity
        try:
            quantity = int(request.POST.get('quantity') or 0)
            if quantity <= 0:
                messages.error(request, "Please enter a valid quantity greater than 0.")
                return redirect(request.META.get('HTTP_REFERER', '/'))
        except (TypeError, ValueError):
            messages.error(request, "Invalid quantity.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Parse price (fallback to stock.share_price if blank)
        price_str = request.POST.get('selling_price')
        try:
            selling_price = (
                Decimal(price_str)
                if price_str not in (None, "",)
                else (stock.share_price or Decimal('0'))
            )
        except (InvalidOperation, TypeError):
            messages.error(request, "Invalid price.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # ✅ Only enforce holdings check for Thangiv (non-Other) advisor
        if (advisor.advisor_type or "").strip().lower() == 'thangiv':
            # Count finalized Thangiv buys across ANY broker
            bought_thangiv = (
                BuyTransaction.objects
                .filter(
                    user=request.user,
                    stock=stock,
                    advisor__advisor_type__iexact='Thangiv',
                    AC_status='completed',           # treat AC-completed as finalized buy
                )
                .aggregate(total=Sum('quantity'))['total'] or 0
            )

            # Count all Thangiv sells except cancelled (includes pending to prevent double-queue oversell)
            sold_thangiv = (
                SellTransaction.objects
                .filter(
                    user=request.user,
                    stock=stock,
                    advisor__advisor_type__iexact='Thangiv',
                )
                .exclude(status='cancelled')
                .aggregate(total=Sum('quantity'))['total'] or 0
            )

            available = max(0, bought_thangiv - sold_thangiv)

            if quantity > available:
                shortfall = quantity - available
                messages.error(
                    request,
                    (
                        f"{stock.company_name}: Requested {quantity} shares, but  {available} available "
                        f"under Thangiv Advisor (shortfall {shortfall}). "
                        f" Please modify the order."
                    )
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))


        # Create the sell order (route based on advisor type)
        if (advisor.advisor_type or "").strip().lower() == 'other':
            SellTransactionOtherAdvisor.objects.create(
                user=request.user,
                stock=stock,
                advisor=advisor,
                broker=broker,
                quantity=quantity,
                selling_price=selling_price,
                status='processing',
            )
        else:
            SellTransaction.objects.create(
                user=request.user,
                stock=stock,
                advisor=advisor,
                broker=broker,
                quantity=quantity,
                selling_price=selling_price,
                status='processing',
            )

        note = _market_closed_note()
        if note:
            messages.info(request, note)
        messages.success(request, f"Sell order placed for {stock.company_name}.")

    except Exception as e:
        messages.error(request, f"Failed to place sell order: {e}")

    return redirect(request.META.get('HTTP_REFERER', '/'))


from Share_Transfer.models import DealLetterRecord

# view_deal_letters



from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .utils import get_user_stock_context  # if you use this elsewhere

@login_required
def view_deal_letters(request):
    user = request.user
    stock_context = {}

    # Optional helper, safe if missing
    try:
        stock_context = get_user_stock_context(user, request)
    except Exception:
        pass

    # Filters
    raw_deal_type = request.GET.get('deal_type', '').strip()  # may be "", "BUY", "SELL" (or lower)
    sort_key = request.GET.get('sort_by', 'date_desc')        # one of: date_desc, date_asc, amount_desc, amount_asc

    records = DealLetterRecord.objects.filter(user=user)

    # Be lenient: accept either upper/lower values
    if raw_deal_type:
        # If your model stores uppercase choice values ("BUY"/"SELL"), __iexact keeps it robust.
        records = records.filter(deal_type__iexact=raw_deal_type)

    # Apply sorting (correct directions)
    allowed_sorts = {
        'date_desc':   '-generated_on',  # newest first
        'date_asc':     'generated_on',  # oldest first
        'amount_desc': '-total_amount',  # high -> low
        'amount_asc':   'total_amount',  # low -> high
    }
    records = records.order_by(allowed_sorts.get(sort_key, '-generated_on'))

    # Normalize selected values for the template
    selected_deal_type = raw_deal_type.upper() if raw_deal_type else ''
    selected_sort = sort_key

    return render(request, 'portfolio/view_deal_letters.html', {
        'records': records,
        'selected_deal_type': selected_deal_type,  # "BUY"/"SELL" or ""
        'selected_sort': selected_sort,            # e.g. "date_desc"
        **stock_context
    })



# sidebar live filter
from django.http import JsonResponse
from django.template.loader import render_to_string
from user_portfolio.utils import get_user_stock_context

def ajax_filtered_stocks(request):
    context = get_user_stock_context(request.user, request)
    html = render_to_string("accounts/includes/sidebar_stock_list.html", context, request=request)
    return JsonResponse({"html": html})
